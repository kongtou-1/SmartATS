import os
import hashlib
import tempfile
import unittest
import uuid
from urllib.parse import urlparse
from datetime import datetime, timedelta, timezone
from io import BytesIO

_TEMP = tempfile.TemporaryDirectory()
os.environ["DATABASE_URL"] = f"sqlite:///{os.path.join(_TEMP.name, 'v2.db').replace(os.sep, '/')}"
os.environ["STORAGE_BACKEND"] = "local"
os.environ["UPLOAD_DIR"] = os.path.join(_TEMP.name, "uploads")

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from app import models
from app.services.database import Base, SessionLocal, engine
from app.main import app
from app.seed import seed_if_empty
from app.services.recruitment import applications, interviews
from app.workers.data_tasks import process_data_job
from app.storage import storage
from app.services.documents.excel_io import build_template, build_export, parse_import
from app.services.documents.offer_pdf import generate_offer_pdf


class V2Test(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        applications.send_email_task.delay = lambda *a, **k: None
        applications.analyze_application_task.delay = lambda *a, **k: None
        interviews.send_email_task.delay = lambda *a, **k: None
        cls.ctx = TestClient(app); cls.client = cls.ctx.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.ctx.__exit__(None, None, None); engine.dispose(); _TEMP.cleanup()

    def setUp(self):
        Base.metadata.drop_all(engine); Base.metadata.create_all(engine); seed_if_empty()

    def auth(self, email):
        r = self.client.post("/api/v1/auth/login", json={"email": email, "password": "demo1234"})
        return {"Authorization": f"Bearer {r.json()['access_token']}"}

    def test_talent_search_and_audit_masking(self):
        hr = self.auth("hr@demo.com")
        channels = self.client.get("/api/v1/admin/source-channels", headers=hr).json()
        data = {"name":"李雷","phone":"+86 139-0000-0000","contact_email":"lilei@example.com","city":"上海","years_experience":6,"source_channel_id":channels[1]["id"],"skills":["Python","SQL"],"tag_ids":[]}
        created = self.client.post("/api/v1/admin/talents", json=data, headers=hr)
        self.assertEqual(created.status_code, 201, created.text)
        found = self.client.get("/api/v1/admin/talents?phone=13900000000&skills=python&skills=sql&min_years=5", headers=hr).json()
        self.assertEqual(found["total"], 1)
        with SessionLocal() as db:
            audit = db.query(models.AuditLog).filter_by(action="TALENT_CREATE").one()
            self.assertTrue(audit.after_data["phone"].startswith("***")); self.assertNotIn("13900000000", str(audit.after_data))

    def test_export_job_preserves_complete_talent_filter(self):
        hr = self.auth("hr@demo.com")
        channels = self.client.get("/api/v1/admin/source-channels", headers=hr).json()
        tags = self.client.post("/api/v1/admin/tags", json={"name":"重点","color":"#2563eb","enabled":True}, headers=hr).json()
        created = self.client.post("/api/v1/admin/talents", json={"name":"导出测试","phone":"13800000001","contact_email":"export@example.com","city":"上海","years_experience":8,"source_channel_id":channels[1]["id"],"skills":["Python","SQL"],"tag_ids":[tags["id"]]}, headers=hr).json()
        response = self.client.post("/api/v1/admin/data-jobs/export", params={"name":"导出测试","phone":"13800000001","skills":"Python,SQL","tag_ids":tags["id"],"source_channel_id":channels[1]["id"],"min_years":5,"max_years":10}, headers=hr)
        self.assertEqual(response.status_code, 202, response.text)
        with SessionLocal() as db:
            row=db.get(models.DataJob,uuid.UUID(response.json()["id"])); filters=row.summary["filters"]
            self.assertEqual(filters["skills"],["Python,SQL"]);self.assertEqual(filters["tag_ids"],[tags["id"]]);self.assertEqual(filters["source_channel_id"],channels[1]["id"]);self.assertEqual(created["name"],filters["name"])

    def test_complete_talent_combination_pagination_and_filtered_export(self):
        hr = self.auth("hr@demo.com")
        source = self.client.get("/api/v1/admin/source-channels", headers=hr).json()[1]
        owner = self.client.get("/api/v1/admin/talent-owners", headers=hr).json()[0]
        tag_a = self.client.post("/api/v1/admin/tags", json={"name":"核心","color":"#2563eb","enabled":True}, headers=hr).json()
        tag_b = self.client.post("/api/v1/admin/tags", json={"name":"可到岗","color":"#16a34a","enabled":True}, headers=hr).json()
        target = self.client.post("/api/v1/admin/talents", json={"name":"组合筛选目标","phone":"+86 138-1234-5678","contact_email":"target@example.com","city":"上海","years_experience":7,"source_channel_id":source["id"],"owner_id":owner["id"],"skills":["Python","SQL"],"tag_ids":[tag_a["id"],tag_b["id"]]}, headers=hr).json()
        self.client.post("/api/v1/admin/talents", json={"name":"组合筛选排除","phone":"13812345679","contact_email":"excluded@example.com","city":"上海","years_experience":7,"source_channel_id":source["id"],"owner_id":owner["id"],"skills":["Python"],"tag_ids":[tag_a["id"]]}, headers=hr)
        with SessionLocal() as db:
            job=db.query(models.Job).filter_by(status="PUBLISHED").first();resume=models.Resume(candidate_id=uuid.UUID(target["id"]),file_name="filter",storage_key="filter",parse_status="DONE",parsed_data={});db.add(resume);db.flush();application=models.Application(candidate_id=uuid.UUID(target["id"]),job_id=job.id,resume_id=resume.id,current_stage="FINAL_REVIEW",owner_id=uuid.UUID(owner["id"]),source_channel_id=uuid.UUID(source["id"]),source_code_snapshot=source["code"]);db.add(application);db.commit();job_id=str(job.id)
        params=[("name","组合筛选"),("phone","13812345678"),("skills","Python,SQL"),("tag_ids",f"{tag_a['id']},{tag_b['id']}"),("source_channel_id",source["id"]),("owner_id",owner["id"]),("job_id",job_id),("stage","FINAL_REVIEW"),("min_years","5"),("max_years","8"),("page","1"),("page_size","1")]
        found=self.client.get("/api/v1/admin/talents",params=params,headers=hr);self.assertEqual(found.status_code,200,found.text);self.assertEqual(found.json()["total"],1);self.assertEqual(found.json()["items"][0]["id"],target["id"]);self.assertEqual(found.json()["page_size"],1)
        queued=self.client.post("/api/v1/admin/data-jobs/export",params=params[:-2],headers=hr);self.assertEqual(queued.status_code,202,queued.text);job_id_export=queued.json()["id"];process_data_job.run(job_id_export)
        with SessionLocal() as db: data_job=db.get(models.DataJob,uuid.UUID(job_id_export));self.assertEqual(data_job.summary["exported"],1);raw=storage.read_file(data_job.output_key)
        workbook=load_workbook(BytesIO(raw),data_only=True);self.assertEqual(workbook.active.max_row,2);self.assertEqual(workbook.active["A2"].value,"组合筛选目标")

    def test_bulk_partial_success_and_idempotency(self):
        hr = self.auth("hr@demo.com")
        with SessionLocal() as db:
            user = db.query(models.CandidateAccount).filter_by(email="candidate@demo.com").one(); c = db.query(models.Candidate).filter_by(user_id=user.id).one(); job=db.query(models.Job).filter_by(status="PUBLISHED").first()
            resume=models.Resume(candidate_id=c.id,file_name="x",storage_key="x",parse_status="DONE",parsed_data={});db.add(resume);db.flush()
            a=models.Application(candidate_id=c.id,job_id=job.id,resume_id=resume.id);db.add(a);db.commit();aid=str(a.id)
        payload={"application_ids":[aid,"00000000-0000-0000-0000-000000000001"],"action":"ADVANCE","reason":"批量筛选","idempotency_key":"bulk-test-0001"}
        r=self.client.post("/api/v1/admin/applications/bulk-actions",json=payload,headers=hr)
        self.assertEqual((r.json()["success_count"],r.json()["failure_count"]),(1,1))
        r2=self.client.post("/api/v1/admin/applications/bulk-actions",json=payload,headers=hr)
        self.assertEqual(r.json(),r2.json())

    def test_excel_roundtrip_and_formula_guard(self):
        raw=build_template(); rows=parse_import(raw); self.assertEqual(rows[0]["姓名"],"示例候选人")
        exported=build_export([["=2+2","138","a@b.com","上海",3,"SQL","高潜","REFERRAL","hr@demo.com","", "岗位","SCREENING",datetime.now(timezone.utc)]])
        wb=load_workbook(BytesIO(exported),data_only=False); self.assertEqual(wb.active["A2"].value,"'=2+2");self.assertIsNone(wb.active["M2"].value.tzinfo);self.assertEqual(wb.active["M2"].number_format,"yyyy-mm-dd hh:mm")

    def test_offer_pdf_and_acceptance(self):
        raw=generate_offer_pdf({"candidate_name":"张三","job_title":"后端工程师","work_location":"上海","salary_description":"月薪 20K","expected_start_date":"2026-09-01","expires_at":"2026-08-30","version":1})
        reader=PdfReader(BytesIO(raw)); self.assertEqual(len(reader.pages),1); self.assertIn("录用通知书",reader.pages[0].extract_text())
        hr, admin, candidate = self.auth("hr@demo.com"), self.auth("admin@demo.com"), self.auth("candidate@demo.com")
        with SessionLocal() as db:
            cu=db.query(models.CandidateAccount).filter_by(email="candidate@demo.com").one(); c=db.query(models.Candidate).filter_by(user_id=cu.id).one(); job=db.query(models.Job).filter_by(status="PUBLISHED").first(); resume=models.Resume(candidate_id=c.id,file_name="x",storage_key="x",parse_status="DONE",parsed_data={});db.add(resume);db.flush(); application=models.Application(candidate_id=c.id,job_id=job.id,resume_id=resume.id,current_stage="FINAL_REVIEW");db.add(application);db.commit(); app_id=str(application.id)
        payload={"application_id":app_id,"salary_description":"20K","work_location":"上海","expected_start_date":(datetime.now(timezone.utc)+timedelta(days=30)).isoformat(),"expires_at":(datetime.now(timezone.utc)+timedelta(days=7)).isoformat(),"probation":"3个月","extra_terms":""}
        offer=self.client.post("/api/v1/admin/offers",json=payload,headers=hr).json(); self.client.post(f"/api/v1/admin/offers/{offer['id']}/submit",json={},headers=hr); self.client.post(f"/api/v1/admin/offers/{offer['id']}/approve",json={"comment":"同意"},headers=admin); self.client.post(f"/api/v1/admin/offers/{offer['id']}/send",json={},headers=hr)
        accepted=self.client.post(f"/api/v1/candidate/offers/{offer['id']}/respond",json={"decision":"ACCEPT","reason":""},headers=candidate)
        self.assertEqual(accepted.json()["status"],"ACCEPTED")
        with SessionLocal() as db: self.assertEqual(db.get(models.Application, uuid.UUID(app_id)).status,"HIRED")

    def test_calendar_manual_busy_blocks_availability_and_delete(self):
        hr = self.auth("hr@demo.com")
        interviewer = self.client.get("/api/v1/admin/interviewers", headers=hr).json()[0]
        tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).replace(hour=2, minute=0, second=0, microsecond=0)
        payload = {"interviewer_id": interviewer["id"], "title": "部门会议", "starts_at": tomorrow.isoformat(), "ends_at": (tomorrow + timedelta(hours=2)).isoformat()}
        created = self.client.post("/api/v1/admin/calendar/busy-blocks", json=payload, headers=hr)
        self.assertEqual(created.status_code, 201, created.text)
        path = f"/api/v1/admin/calendar/interviewers/{interviewer['id']}/availability"
        params = {"start": tomorrow.isoformat(), "end": (tomorrow + timedelta(hours=3)).isoformat(), "duration_minutes": 60}
        blocked = self.client.get(path, params=params, headers=hr).json()["slots"]
        self.assertTrue(all(datetime.fromisoformat(slot["start"]) >= tomorrow + timedelta(hours=2) for slot in blocked))
        deleted = self.client.delete(f"/api/v1/admin/calendar/busy-blocks/{created.json()['id']}", headers=hr)
        self.assertEqual(deleted.status_code, 204)
        restored = self.client.get(path, params=params, headers=hr).json()["slots"]
        self.assertGreater(len(restored), len(blocked))

    def test_public_offer_link_pdf_response_and_one_time_token(self):
        hr, admin = self.auth("hr@demo.com"), self.auth("admin@demo.com")
        with SessionLocal() as db:
            cu=db.query(models.CandidateAccount).filter_by(email="candidate@demo.com").one(); c=db.query(models.Candidate).filter_by(user_id=cu.id).one(); job=db.query(models.Job).filter_by(status="PUBLISHED").first(); resume=models.Resume(candidate_id=c.id,file_name="public",storage_key="public",parse_status="DONE",parsed_data={});db.add(resume);db.flush(); application=models.Application(candidate_id=c.id,job_id=job.id,resume_id=resume.id,current_stage="FINAL_REVIEW");db.add(application);db.commit(); app_id=str(application.id)
        payload={"application_id":app_id,"salary_description":"25K","work_location":"上海","expected_start_date":(datetime.now(timezone.utc)+timedelta(days=30)).isoformat(),"expires_at":(datetime.now(timezone.utc)+timedelta(days=7)).isoformat(),"probation":"3个月","extra_terms":""}
        offer=self.client.post("/api/v1/admin/offers",json=payload,headers=hr).json();self.client.post(f"/api/v1/admin/offers/{offer['id']}/submit",json={},headers=hr);self.client.post(f"/api/v1/admin/offers/{offer['id']}/approve",json={"comment":"同意"},headers=admin)
        sent=self.client.post(f"/api/v1/admin/offers/{offer['id']}/send",json={},headers=hr).json(); token=urlparse(sent["simulated_response_url"]).path.rsplit("/",1)[-1]
        self.assertEqual(self.client.get(f"/api/v1/offers/respond/{token}").status_code,200)
        pdf=self.client.get(f"/api/v1/offers/respond/{token}/pdf");self.assertEqual(pdf.status_code,200);self.assertTrue(pdf.content.startswith(b"%PDF"))
        accepted=self.client.post(f"/api/v1/offers/respond/{token}",json={"decision":"ACCEPT","reason":""});self.assertEqual(accepted.json()["status"],"ACCEPTED")
        self.assertEqual(self.client.get(f"/api/v1/offers/respond/{token}").status_code,404)

    def test_public_offer_decline_rejects_application_and_invalidates_token(self):
        raw_token="decline-once-token";digest=hashlib.sha256(raw_token.encode()).hexdigest()
        with SessionLocal() as db:
            creator=db.query(models.AdminAccount).filter_by(email="hr@demo.com").one();candidate=db.query(models.Candidate).first();job=db.query(models.Job).filter_by(status="PUBLISHED").first();resume=models.Resume(candidate_id=candidate.id,file_name="decline",storage_key="decline",parse_status="DONE",parsed_data={});db.add(resume);db.flush();application=models.Application(candidate_id=candidate.id,job_id=job.id,resume_id=resume.id,current_stage="FINAL_REVIEW");db.add(application);db.flush();offer=models.Offer(application_id=application.id,candidate_id=candidate.id,job_id=job.id,status="SENT",salary_description="20K",work_location="上海",expected_start_date=datetime.now(timezone.utc)+timedelta(days=30),expires_at=datetime.now(timezone.utc)+timedelta(days=7),created_by=creator.id,response_token_hash=digest);db.add(offer);db.commit();app_id=application.id
        declined=self.client.post(f"/api/v1/offers/respond/{raw_token}",json={"decision":"DECLINE","reason":"接受了其他机会"});self.assertEqual(declined.status_code,200,declined.text);self.assertEqual(declined.json()["status"],"DECLINED")
        self.assertEqual(self.client.post(f"/api/v1/offers/respond/{raw_token}",json={"decision":"DECLINE","reason":"重复"}).status_code,404)
        with SessionLocal() as db: application=db.get(models.Application,app_id);self.assertEqual(application.status,"REJECTED");self.assertEqual(application.current_stage,"REJECTED")


if __name__ == "__main__": unittest.main()
