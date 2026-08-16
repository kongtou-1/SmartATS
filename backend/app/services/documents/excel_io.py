"""Spreadsheet document generation and parsing."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["姓名", "手机号", "邮箱", "城市", "工作年限", "技能", "标签", "来源编码", "负责人邮箱", "备注"]
EXPORT_HEADERS = HEADERS + ["最近岗位", "最近阶段", "创建时间"]


def safe_excel(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(ZoneInfo("Asia/Shanghai")).replace(tzinfo=None)
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _style(ws, width_count):
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{chr(64 + width_count)}1" if width_count <= 26 else ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="16324F"); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
    widths = [16, 16, 28, 14, 12, 28, 22, 16, 28, 32, 22, 16, 20]
    for i in range(1, width_count + 1): ws.column_dimensions[chr(64 + i)].width = widths[i - 1]


def build_template() -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "人才导入"
    ws.append(HEADERS); ws.append(["示例候选人", "13800000000", "name@example.com", "上海", 5, "Python,SQL", "高潜,后端", "REFERRAL", "hr@demo.com", "示例行可删除"])
    _style(ws, len(HEADERS)); notes = wb.create_sheet("填写说明")
    notes.append(["字段", "说明"])
    for row in [("必填", "姓名，以及手机号/邮箱至少一个"), ("技能/标签", "使用英文逗号分隔"), ("重复规则", "手机号优先、邮箱次之；匹配后合并非空资料"), ("来源编码", "UNKNOWN/CAREER_SITE/REFERRAL/JOB_BOARD/AGENCY/CAMPUS/OTHER")]: notes.append(row)
    _style(notes, 2); stream = BytesIO(); wb.save(stream); return stream.getvalue()


def parse_import(raw: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True); ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows()
    headers = [str(c.value or "").strip() for c in next(iterator)]
    missing = [h for h in HEADERS if h not in headers]
    if missing: raise ValueError(f"缺少列：{','.join(missing)}")
    index = {h: headers.index(h) for h in HEADERS}; output = []
    for row_no, cells in enumerate(iterator, start=2):
        values = [c.value if c.value is not None else "" for c in cells]
        if not any(values): continue
        output.append({h: values[i] if i < len(values) else "" for h, i in index.items()} | {"_row": row_no})
    return output


def build_export(rows: list[list]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "人才库"
    ws.append(EXPORT_HEADERS)
    for row in rows: ws.append([safe_excel(v) for v in row])
    _style(ws, len(EXPORT_HEADERS)); ws.column_dimensions["M"].width = 20
    for cell in ws["M"][1:]: cell.number_format = "yyyy-mm-dd hh:mm"
    stream = BytesIO(); wb.save(stream); return stream.getvalue()


def build_import_result(errors: list[dict]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "导入错误"
    ws.append(["行号", "错误"])
    for item in errors: ws.append([item.get("row"), safe_excel(item.get("error", ""))])
    _style(ws, 2); stream = BytesIO(); wb.save(stream); return stream.getvalue()


def build_table(sheet_name: str, headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = sheet_name[:31]
    ws.append(headers)
    for row in rows: ws.append([safe_excel(v) for v in row])
    _style(ws, len(headers)); stream = BytesIO(); wb.save(stream); return stream.getvalue()
